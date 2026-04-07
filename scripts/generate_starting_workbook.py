"""Generate PR_Simulator starting builds Excel. Run: python scripts/generate_starting_workbook.py"""
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "data" / "starting_builds_and_economy.xlsx"


def total_v(eur: float, comp: float, vis: float, cap: float) -> float:
    return eur / 80000 + comp / 80 + vis / 80 + cap / 50


def main() -> None:
    wb = Workbook()

    # --- Sheet 1: Starting builds ---
    ws1 = wb.active
    ws1.title = "Starting_Builds"
    headers1 = [
        "Build ID",
        "Display name",
        "Starting EUR",
        "Competence",
        "Visibility",
        "Firm capacity",
        "Total_V",
        "Opening hire rule (design)",
    ]
    rows1 = [
        [
            "velvet_rolodex",
            "The Velvet Rolodex",
            16000,
            30,
            80,
            50,
            total_v(16000, 30, 80, 50),
            "Two junior employees (salary bands 16k–30k EUR / year)",
        ],
        [
            "summa_cum_basement",
            "Summa Cum Basement",
            16000,
            80,
            30,
            50,
            total_v(16000, 80, 30, 50),
            "Two junior employees (salary bands 16k–30k EUR / year)",
        ],
        [
            "portfolio_pivot",
            "The Portfolio Pivot",
            80000,
            22,
            24,
            50,
            total_v(80000, 22, 24, 50),
            "One mid-tier at top of band (up to 70k EUR / year class)",
        ],
    ]
    _write_table(ws1, headers1, rows1)

    # --- Sheet 2: Salary ranges ---
    ws2 = wb.create_sheet("Salary_Ranges")
    headers2 = ["Tier", "Annual salary (EUR)", "Notes"]
    rows2 = [
        ["Junior", "16,000 – 30,000", "Low builds start with two juniors"],
        ["Mid", "31,000 – 70,000", "Portfolio Pivot opens with best mid-tier"],
        ["Senior", "71,000 – 100,000", "Unlocked later / higher payroll"],
        ["Executive", "100,000+", "Late game"],
    ]
    _write_table(ws2, headers2, rows2)

    # --- Sheet 3: Benchmark conversion ---
    ws3 = wb.create_sheet("Benchmark_Conversion")
    ws3["A1"] = "Calibration premise"
    ws3["A1"].font = Font(bold=True)
    ws3["A2"] = (
        "At benchmark, these represent equal in-game value: "
        "80,000 EUR ≙ 80 competence ≙ 80 visibility ≙ 50 firm capacity."
    )
    ws3.merge_cells("A2:F2")
    ws3["A2"].alignment = Alignment(wrap_text=True)

    headers3 = [
        "Conversion",
        "Formula / value",
        "Detail",
    ]
    rows3 = [
        [
            "EUR per 1 competence point",
            80000 / 80,
            "1000 EUR per point at benchmark line",
        ],
        [
            "EUR per 1 visibility point",
            80000 / 80,
            "1000 EUR per point (same weight as competence at bench)",
        ],
        [
            "EUR per 1 firm capacity point",
            80000 / 50,
            "1600 EUR per point (scarcer per unit than comp/vis)",
        ],
        [
            "1 competence point ≡ capacity points (same value V)",
            50 / 80,
            "0.625 capacity points (capacity is fewer points for same V)",
        ],
        [
            "1 firm capacity point ≡ competence points (same value V)",
            80 / 50,
            "1.6 competence points",
        ],
        [
            "Full reference",
            "data/conversion_mechanics.txt",
            "Backend design document",
        ],
    ]
    _write_table(ws3, headers3, rows3, start_row=4)

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"Wrote {OUT}")


def _write_table(ws, headers, rows, start_row: int = 1) -> None:
    r = start_row
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=r, column=c, value=h)
        cell.font = Font(bold=True)
    r += 1
    for row in rows:
        for c, val in enumerate(row, 1):
            ws.cell(row=r, column=c, value=val)
        r += 1
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 28


if __name__ == "__main__":
    main()
